#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""F1 — Motore locale numeri/vie per microzone.

Legge da GitHub SOLO le vie (nessun dato personale), cerca profili pubblici
PagineBianche/PagineGialle per le microzone correnti e salva i numeri SOLO sul PC.
Genera LISTA_MATTINO.html per il lavoro telefonico dalla scrivania.

Non aggira CAPTCHA. Se un motore di ricerca mostra un controllo, lo registra e
prova gli altri provider; Chrome rimane visibile.
"""
from __future__ import annotations

import csv
import html
import json
import os
import random
import re
import sqlite3
import subprocess
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

BASE = Path.home() / "Documents" / "F1_Directory_Microzone"
DATA = BASE / "data"
IMPORT_DIR = BASE / "IMPORTA_ESISTENTI"
DB = DATA / "contacts.sqlite"
TARGETS_CSV = DATA / "microzone_targets.csv"
OUT_CSV = DATA / "telefonate_mattino.csv"
OUT_HTML = BASE / "LISTA_MATTINO.html"
CAPTCHA_STATE = DATA / "captcha_state.json"
LOG = BASE / "motore.log"

TARGETS_URL = "https://raw.githubusercontent.com/josephsocialmedia2-spec/immobili-in-zona/main/seller_radar_auto/data/microzone_targets.csv"
MAIN_TOWNS = {
    "avigliana", "almese", "condove", "sant antonino di susa",
    "vaie", "chiusa di san michele", "bussoleno", "susa"
}
PROVIDERS = ["google", "bing", "duckduckgo"]
MAX_RESULT_PAGES = 3
SEARCH_DELAY = (7.0, 12.0)
PROFILE_DELAY = (2.5, 5.0)


def log(msg):
    BASE.mkdir(parents=True, exist_ok=True)
    line = f"{datetime.now():%Y-%m-%d %H:%M:%S} | {msg}"
    print(line)
    with LOG.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    s = s.replace("’", "'")
    s = re.sub(r"[^a-z0-9' ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def digits(s):
    return re.sub(r"\D+", "", str(s or ""))


def ensure():
    DATA.mkdir(parents=True, exist_ok=True)
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB)
    conn.execute("""
      create table if not exists contacts(
        id integer primary key autoincrement,
        name text, phone text, street text, civic text, comune text,
        source text, source_url text, fetched_at text,
        unique(phone, street, comune)
      )
    """)
    conn.commit(); conn.close()


def download_targets():
    req = urllib.request.Request(TARGETS_URL, headers={"User-Agent":"F1Immobiliare-Desktop/1.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        raw = r.read()
    TARGETS_CSV.write_bytes(raw)
    log(f"Microzone aggiornate: {len(raw)} byte")


def read_targets():
    if not TARGETS_CSV.exists():
        return []
    with TARGETS_CSV.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    return [r for r in rows if norm(r.get("COMUNE")) in MAIN_TOWNS and r.get("VIA_DA_LAVORARE")]


def upsert_contact(c):
    p = digits(c.get("phone"))
    if len(p) < 7:
        return False
    conn = sqlite3.connect(DB)
    conn.execute("""
      insert into contacts(name,phone,street,civic,comune,source,source_url,fetched_at)
      values(?,?,?,?,?,?,?,?)
      on conflict(phone,street,comune) do update set
        name=excluded.name, civic=excluded.civic, source=excluded.source,
        source_url=excluded.source_url, fetched_at=excluded.fetched_at
    """, (
        c.get("name", ""), c.get("phone", ""), c.get("street", ""),
        c.get("civic", ""), c.get("comune", ""), c.get("source", ""),
        c.get("source_url", ""), datetime.now().isoformat(timespec="seconds")
    ))
    conn.commit(); conn.close(); return True


def import_csv(path):
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    try:
        delim = csv.Sniffer().sniff(raw[:4096], delimiters=";,\t,").delimiter
    except Exception:
        delim = ";"
    n = 0
    for r in csv.DictReader(raw.splitlines(), delimiter=delim):
        low = {norm(k): v for k,v in r.items()}
        phone = low.get("telefono") or low.get("phone") or low.get("tel") or ""
        street = low.get("via") or low.get("indirizzo") or low.get("address") or ""
        comune = low.get("comune") or ""
        civic = low.get("civico") or ""
        name = low.get("nome") or low.get("nominativo") or low.get("name") or ""
        if digits(phone) and street and norm(comune) in MAIN_TOWNS:
            if upsert_contact({"name":name,"phone":phone,"street":street,"civic":civic,"comune":comune,"source":low.get("fonte") or path.name,"source_url":low.get("link") or low.get("url") or ""}):
                n += 1
    return n


def import_xlsx(path):
    try:
        from openpyxl import load_workbook
    except Exception:
        return 0
    wb = load_workbook(path, read_only=True, data_only=True)
    n = 0
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True); headers = None
        for vals in rows:
            vals = ["" if v is None else str(v) for v in vals]
            nvals = [norm(v) for v in vals]
            if headers is None:
                if "telefono" in nvals and ("via" in nvals or "indirizzo" in nvals):
                    headers = vals
                continue
            r = {headers[i]: vals[i] if i < len(vals) else "" for i in range(len(headers))}
            low = {norm(k):v for k,v in r.items()}
            phone = low.get("telefono") or low.get("phone") or ""
            street = low.get("via") or low.get("indirizzo") or ""
            comune = low.get("comune") or ""
            if digits(phone) and street and norm(comune) in MAIN_TOWNS:
                if upsert_contact({"name":low.get("nome") or low.get("nominativo") or "","phone":phone,"street":street,"civic":low.get("civico") or "","comune":comune,"source":low.get("fonte") or path.name,"source_url":low.get("link") or low.get("url") or ""}):
                    n += 1
    return n


def import_existing():
    total = 0
    candidates = list(IMPORT_DIR.glob("*.csv")) + list(IMPORT_DIR.glob("*.xlsx"))
    # Importa anche gli export già prodotti dal bridge F1, se presenti.
    bridge = Path.home()/"Documents"/"F1_Bridge"/"IMPORTA_CONTATTI"
    if bridge.exists():
        candidates += list(bridge.glob("*.csv")) + list(bridge.glob("*.xlsx"))
    for p in candidates:
        try:
            total += import_csv(p) if p.suffix.lower()==".csv" else import_xlsx(p)
        except Exception as e:
            log(f"Import {p.name}: {e}")
    if total:
        log(f"Contatti locali importati/aggiornati: {total}")


def notify_captcha(provider, url):
    state = {"status":"ATTESA_OPERATORE","provider":provider,"url":url,"detected_at":datetime.now().isoformat(timespec="seconds")}
    CAPTCHA_STATE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        msg = f"CAPTCHA rilevato su {provider}. Chrome resta aperto. Il motore prova le altre fonti."
        ps = "Add-Type -AssemblyName System.Windows.Forms;[System.Windows.Forms.MessageBox]::Show('"+msg.replace("'","''")+"','F1 - CAPTCHA',[System.Windows.Forms.MessageBoxButtons]::OK,[System.Windows.Forms.MessageBoxIcon]::Warning)|Out-Null"
        subprocess.Popen(["powershell.exe","-NoProfile","-Command",ps])
    except Exception:
        pass


def captcha_present(driver):
    try:
        text = ((driver.title or "") + " " + (driver.page_source or "")).lower()
    except Exception:
        return False
    return any(x in text for x in ["captcha","recaptcha","unusual traffic","verify you are human","verifica di essere umano","cloudflare"])


def make_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    profile = BASE / "chrome_profile"
    profile.mkdir(parents=True, exist_ok=True)
    o = Options()
    o.add_argument(f"--user-data-dir={profile}")
    o.add_argument("--start-maximized")
    o.add_argument("--disable-blink-features=AutomationControlled")
    d = webdriver.Chrome(options=o)
    d.set_page_load_timeout(35)
    return d


def search_url(provider, q, start=0):
    qq = urllib.parse.quote_plus(q)
    if provider == "google":
        return f"https://www.google.com/search?q={qq}&start={start}"
    if provider == "bing":
        return f"https://www.bing.com/search?q={qq}&first={start+1}"
    return f"https://html.duckduckgo.com/html/?q={qq}&s={start}"


def extract_profile_links(driver, domain):
    links = []
    for a in driver.find_elements("css selector", "a[href]"):
        try:
            u = a.get_attribute("href") or ""
        except Exception:
            continue
        lu = u.lower()
        if domain == "paginebianche" and "paginebianche.it/" in lu and ("/scheda/" in lu or ".htm" not in lu):
            links.append(u)
        if domain == "paginegialle" and "paginegialle.it/" in lu and "/piemonte/" not in lu:
            links.append(u)
    # canonicalizza e limita
    out=[]; seen=set()
    for u in links:
        u=u.split("#")[0]
        if u not in seen:
            seen.add(u); out.append(u)
    return out[:50]


def parse_profile(driver, target):
    try:
        body = driver.find_element("tag name", "body").text
    except Exception:
        body = ""
    src = driver.page_source or ""
    if norm(target["VIA_DA_LAVORARE"]) not in norm(body + " " + src):
        return []
    if norm(target["COMUNE"]) not in norm(body + " " + src):
        return []
    try:
        name = driver.find_element("css selector", "h1").text.strip()
    except Exception:
        name = (driver.title or "").split("|")[0].strip()
    phones = set()
    for m in re.findall(r'href=["\']tel:([^"\']+)', src, flags=re.I):
        p = digits(urllib.parse.unquote(m))
        if 7 <= len(p) <= 13:
            phones.add(p)
    if not phones:
        for m in re.findall(r"(?<!\d)(?:\+39\s*)?(?:0\d{1,3}[\s.-]?\d{5,8}|3\d{2}[\s.-]?\d{6,7})(?!\d)", body):
            p=digits(m)
            if 7 <= len(p) <= 13:
                phones.add(p)
    civic = ""
    sm = re.search(re.escape(target["VIA_DA_LAVORARE"])+r"\s*[, ]+\s*(\d+[A-Za-z]?(?:/\d+[A-Za-z]?)?)", body, flags=re.I)
    if sm:
        civic = sm.group(1)
    source = "PagineBianche" if "paginebianche.it" in driver.current_url.lower() else "PagineGialle"
    return [{"name":name,"phone":p,"street":target["VIA_DA_LAVORARE"],"civic":civic,"comune":target["COMUNE"],"source":source,"source_url":driver.current_url} for p in phones]


def collect_target(driver, target):
    found = 0
    queries = [
      ("paginebianche", f'site:paginebianche.it "{target["VIA_DA_LAVORARE"]}" "{target["COMUNE"]}"'),
      ("paginegialle", f'site:paginegialle.it "{target["VIA_DA_LAVORARE"]}" "{target["COMUNE"]}"'),
    ]
    for domain, q in queries:
        profile_urls=[]
        for provider in PROVIDERS:
            blocked=False
            for page in range(MAX_RESULT_PAGES):
                try:
                    driver.get(search_url(provider,q,page*10))
                    time.sleep(random.uniform(*SEARCH_DELAY))
                    if captcha_present(driver):
                        notify_captcha(provider, driver.current_url); blocked=True; break
                    profile_urls += extract_profile_links(driver, domain)
                except Exception as e:
                    log(f"Ricerca {provider}: {e}"); break
            if profile_urls:
                break
            if blocked:
                continue
        # profili unici
        seen=[]
        for u in profile_urls:
            if u not in seen: seen.append(u)
        for u in seen[:60]:
            try:
                driver.get(u); time.sleep(random.uniform(*PROFILE_DELAY))
                if captcha_present(driver):
                    notify_captcha(domain, driver.current_url); continue
                for c in parse_profile(driver,target):
                    if upsert_contact(c): found += 1
            except Exception as e:
                log(f"Profilo {u[:80]}: {e}")
    return found


def existing_for_target(t):
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row
    rows=conn.execute("select * from contacts where lower(comune)=lower(?)",(t["COMUNE"],)).fetchall(); conn.close()
    sv=norm(t["VIA_DA_LAVORARE"])
    return [dict(r) for r in rows if sv and (sv in norm(r["street"]) or norm(r["street"]) in sv)]


def generate_report(targets):
    rows=[]
    for t in targets:
        for c in existing_for_target(t):
            rows.append({
              "ZONE_ID":t.get("ZONE_ID",""),"COMUNE":t.get("COMUNE",""),
              "VIA_ANNUNCIO":t.get("VIA_ANNUNCIO",""),"RIFERIMENTO_ANNUNCIO":t.get("RIFERIMENTO_ANNUNCIO",""),
              "ANNUNCIO_URL":t.get("ANNUNCIO_URL",""),"RANK_VIA":t.get("RANK",""),
              "VIA_CONTATTO":t.get("VIA_DA_LAVORARE",""),"NOME":c.get("name","") or "",
              "TELEFONO":c.get("phone","") or "","CIVICO":c.get("civic","") or "",
              "FONTE_CONTATTO":c.get("source","") or "","URL_CONTATTO":c.get("source_url","") or "",
              "RPO_STATUS":"DA_VERIFICARE","OWNER_VERIFIED":"NO"
            })
    # deduplica: stesso numero nella stessa zona, priorità via più bassa
    best={}
    for r in rows:
        k=(r["ZONE_ID"],digits(r["TELEFONO"]))
        if not k[1]: continue
        if k not in best or int(r["RANK_VIA"] or 99)<int(best[k]["RANK_VIA"] or 99): best[k]=r
    rows=sorted(best.values(),key=lambda r:(r["COMUNE"],r["VIA_ANNUNCIO"],int(r["RANK_VIA"] or 99),r["VIA_CONTATTO"],r["CIVICO"]))
    fields=["ZONE_ID","COMUNE","VIA_ANNUNCIO","RIFERIMENTO_ANNUNCIO","ANNUNCIO_URL","RANK_VIA","VIA_CONTATTO","NOME","TELEFONO","CIVICO","FONTE_CONTATTO","URL_CONTATTO","RPO_STATUS","OWNER_VERIFIED"]
    with OUT_CSV.open("w",encoding="utf-8-sig",newline="") as f:
        w=csv.DictWriter(f,fieldnames=fields);w.writeheader();w.writerows(rows)

    groups={}
    for r in rows: groups.setdefault((r["COMUNE"],r["VIA_ANNUNCIO"],r["ANNUNCIO_URL"],r["RIFERIMENTO_ANNUNCIO"]),[]).append(r)
    parts=[]
    for (comune,via,url,ref), rs in groups.items():
        trs=[]
        for r in rs:
            trs.append(f"<tr><td>{html.escape(r['VIA_CONTATTO'])}</td><td>{html.escape(r['CIVICO'])}</td><td>{html.escape(r['NOME'])}</td><td><b>{html.escape(r['TELEFONO'])}</b></td><td>{html.escape(r['FONTE_CONTATTO'])}</td><td>DA VERIFICARE</td></tr>")
        parts.append(f"<section class='card'><h2>{html.escape(comune)} — {html.escape(via)}</h2><p>{html.escape(ref)}</p><p><a href='{html.escape(url)}' target='_blank'>APRI ANNUNCIO</a> · {len(rs)} numeri trovati nella microzona</p><table><thead><tr><th>Via</th><th>Civico</th><th>Contatto pubblico</th><th>Telefono</th><th>Fonte</th><th>RPO</th></tr></thead><tbody>{''.join(trs)}</tbody></table></section>")
    page=f"""<!doctype html><html><head><meta charset='utf-8'><title>F1 Lista Mattino</title><style>body{{font-family:Arial;background:#070907;color:white;margin:0}}main{{max-width:1200px;margin:auto;padding:22px}}h1{{color:#39f28a}}.card{{background:#101510;border:1px solid #2a342c;border-radius:16px;padding:16px;margin:14px 0}}table{{width:100%;border-collapse:collapse;font-size:13px}}td,th{{border-bottom:1px solid #2a342c;padding:8px;text-align:left}}a{{color:#39f28a}}p{{color:#aeb7b0}}.note{{background:#17140d;border:1px solid #5b4c25;padding:12px;border-radius:12px}}</style></head><body><main><h1>F1 — LISTA MATTINO</h1><p>Generata {datetime.now():%d/%m/%Y %H:%M}. Annuncio → via centrale → massimo 4 vie vicine → contatti pubblici trovati.</p><div class='note'><b>IMPORTANTE:</b> stesso civico/stessa via non significa proprietario. Prima di una telefonata commerciale verifica RPO e base giuridica applicabile.</div>{''.join(parts) if parts else '<div class="card">Nessun contatto trovato nelle microzone correnti. Controlla motore.log e CAPTCHA.</div>'}</main></body></html>"""
    OUT_HTML.write_text(page,encoding="utf-8")
    log(f"Lista mattino: {len(rows)} numeri, {len(groups)} microzone")
    return rows


def main():
    ensure()
    try: download_targets()
    except Exception as e: log(f"Download microzone: {e}")
    targets=read_targets()
    if not targets:
        log("Nessuna microzona disponibile. Il Radar deve generare microzone_targets.csv."); return 2
    import_existing()
    # Prima genera con ciò che è già in archivio; poi arricchisce le vie che hanno pochi contatti.
    generate_report(targets)
    driver=None
    try:
        driver=make_driver()
        unique=[]; seen=set()
        for t in targets:
            k=(norm(t["COMUNE"]),norm(t["VIA_DA_LAVORARE"]))
            if k not in seen:
                seen.add(k); unique.append(t)
        for i,t in enumerate(unique,1):
            existing=len(existing_for_target(t))
            if existing>=8:
                log(f"[{i}/{len(unique)}] {t['COMUNE']} {t['VIA_DA_LAVORARE']}: cache {existing}, salto ricerca")
                continue
            log(f"[{i}/{len(unique)}] cerco {t['COMUNE']} — {t['VIA_DA_LAVORARE']}")
            n=collect_target(driver,t); log(f"  nuovi/aggiornati: {n}")
        generate_report(targets)
        # Se non ci sono CAPTCHA chiude Chrome. Se ce ne sono, lo lascia aperto sulla sessione visibile.
        if not CAPTCHA_STATE.exists() or json.loads(CAPTCHA_STATE.read_text(encoding="utf-8")).get("status")!="ATTESA_OPERATORE":
            driver.quit(); driver=None
    except Exception as e:
        log(f"Errore motore: {e}")
    finally:
        if driver is not None:
            log("Chrome lasciato aperto: possibile controllo manuale/CAPTCHA.")
    # Apre la lista locale al termine.
    try: os.startfile(OUT_HTML)
    except Exception: pass
    return 0


if __name__=="__main__":
    raise SystemExit(main())
